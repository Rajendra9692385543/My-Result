from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, Response
import os, io, re
from io import BytesIO
import pandas as pd
from supabase import create_client, Client
from flask import make_response, send_file
from reportlab.lib.pagesizes import landscape, A4, letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================
# Supabase Config
# ==========================
SUPABASE_URL = "https://xsnktkaofijftripytij.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inhzbmt0a2FvZmlqZnRyaXB5dGlqIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTI2NDM2MDAsImV4cCI6MjA2ODIxOTYwMH0.NKo7WAeN7ssRg_5LceDABBaUJF-jAEjxBrCjKtoxvgg"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ==========================
# Flask Setup
# ==========================
app = Flask(__name__)
app.secret_key = 'secret_key_here'

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ==========================
# File Serving Route
# ==========================
@app.route('/uploads/<filename>')
def serve_uploaded_file(filename):
    return send_from_directory('uploads', filename, as_attachment=True)

# ==========================
# School → Branch Mapping
# ==========================
school_branch_map = {
    "School of Engineering and Technology": [
        "Computer Science and Engineering",
        "Mechanical Engineering",
        "Electrical and Electronics Engineering",
        "Civil Engineering",
        "Agricultural Engineering",
        "Biotechnology Engineering",
        "Aerospace Engineering",
        "Bachelor of Computer Application (BCA)"
    ],
    "M.S. Swaminathan School of Agriculture": [
        "Agronomy", "Horticulture", "Soil Science"
    ],
    "School of Management": ["BBA", "MBA"],
    "School of Fisheries": ["Fisheries Science"],
    "School of Vocational Education and Training": ["Vocational Training"],
    "School of Applied Sciences": ["Physics", "Chemistry", "Mathematics"],
    "School of Agriculture & Bio-Engineering": ["Bio-Engineering"],
    "School of Veterinary and Animal Sciences": ["Veterinary Science"],
    "School of Nursing": ["Nursing"]
}

# ==========================
# School → Program Mapping
# ==========================
school_program_map = {
    "School of Engineering and Technology": ["BTech", "BCA", "Diploma"],
    "M.S. Swaminathan School of Agriculture": ["BSc AG"],
    "School of Management": ["BBA", "MBA"],
    "School of Fisheries": ["BSc Fisheries"],
    "School of Vocational Education and Training": ["Diploma"],
    "School of Applied Sciences": ["BSc"],
    "School of Agriculture & Bio-Engineering": ["BTech Bio", "BSc Bio"],
    "School of Veterinary and Animal Sciences": ["BVSc"],
    "School of Nursing": ["BSc Nursing"]
}

# ==========================
# Helpers
# ==========================
def calculate_grade_point(grade):
    mapping = {'O': 10, 'E': 9, 'A': 8, 'B': 7, 'C': 6, 'D': 5, 'F': 0, 'S': 0}
    return mapping.get(grade.upper(), 0)

def calculate_gpa(records):
    total_credits = 0
    weighted_points = 0
    for row in records:
        try:
            credits = sum(map(float, row['credits'].split('+'))) if '+' in row['credits'] else float(row['credits'])
            points = calculate_grade_point(row['grade'])
            total_credits += credits
            weighted_points += points * credits
        except:
            continue
    return round(weighted_points / total_credits, 2) if total_credits else 0.0

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==========================
# Routes
# ==========================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/result', methods=['POST'])
def result():
    reg_no = request.form['reg_no'].strip()
    semester = request.form['semester'].strip()

    # Get current semester result including subject_code
    response = supabase.table("results").select("*").eq("reg_no", reg_no).eq("semester", semester).execute()
    semester_records = response.data if response.data else []

    # Get all semesters for CGPA
    response_all = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    all_records = response_all.data if response_all.data else []

    if not semester_records:
        return render_template(
            'result.html',
            name="Not Found",
            reg_no=reg_no,
            semester=semester,
            sgpa=0,
            cgpa=0,
            subjects=[],
            chart_data=[],
            school="-",
            branch="-",
            academic_year="-"
        )

    name = semester_records[0]['name']
    school = semester_records[0]['school']
    branch = semester_records[0]['branch']
    academic_year = semester_records[0]['academic_year']

    sgpa = calculate_gpa(semester_records)
    cgpa = calculate_gpa(all_records)

    # ✅ Get distinct semester list and sort by semester number
    semesters_resp = supabase.table("results").select("semester").eq("reg_no", reg_no).execute()
    raw_semesters = list({row["semester"] for row in semesters_resp.data if row.get("semester")})

    def extract_sem_number(sem):
        try:
            return int(''.join(filter(str.isdigit, sem)))
        except:
            return float('inf')  # push unknowns to end

    semesters = sorted(raw_semesters, key=extract_sem_number)

    # ✅ Prepare chart data in sorted order
    chart_data = []
    for sem in semesters:
        response = supabase.table("results").select("*").eq("reg_no", reg_no).eq("semester", sem).execute()
        rows = response.data
        gpa = calculate_gpa(rows)
        chart_data.append({"semester": sem, "cgpa": gpa})

    return render_template(
        'result.html',
        name=name,
        reg_no=reg_no,
        semester=semester,
        sgpa=sgpa,
        cgpa=cgpa,
        subjects=semester_records,   # now includes subject_code
        chart_data=chart_data,
        school=school,
        branch=branch,
        academic_year=academic_year
    )

@app.route("/download_report/<reg_no>/<semester>")
def download_report(reg_no, semester):
    # ✅ Fetch semester results
    results_resp = supabase.table("results").select("*").eq("reg_no", reg_no).eq("semester", semester).execute()
    results = results_resp.data

    if not results:
        return "No data available", 400

    # ✅ Fetch all semesters for CGPA
    all_resp = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    all_records = all_resp.data if all_resp.data else []

    # ✅ Extract details
    student = {
        "reg_no": results[0]["reg_no"],
        "name": results[0].get("name", "N/A"),
        "school": results[0].get("school", "N/A"),
        "branch": results[0].get("branch", "N/A"),
        "semester": semester,
        "academic_year": results[0].get("academic_year", "N/A"),
    }

    # ✅ Calculate SGPA & CGPA
    sgpa = calculate_gpa(results)
    cgpa = calculate_gpa(all_records)

    # ✅ Generate PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph("<b>Report Card</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Student Info Table (extended)
    student_info = [
        ["Name", student["name"]],
        ["Registration No", student["reg_no"]],
        ["School", student["school"]],
        ["Branch", student["branch"]],
        ["Semester", student["semester"]],
        ["Academic Year", student["academic_year"]],
        ["SGPA", f"{sgpa:.2f}"],
        ["CGPA", f"{cgpa:.2f}"],
    ]
    student_table = Table(student_info, colWidths=[120, 350])
    student_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
    ]))
    elements.append(student_table)
    elements.append(Spacer(1, 14))

    # Subject Results Table
    data = [["Subject Code", "Subject Name", "Grade", "Credits"]]
    for r in results:
        data.append([
            r["subject_code"],
            Paragraph(r["subject_name"], styles["Normal"]),
            r["grade"],
            str(r["credits"])
        ])

    table = Table(data, colWidths=[80, 250, 60, 60])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Report_{reg_no}_Sem{semester}.pdf",
        mimetype="application/pdf"
    )

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()

        if not email or not password:
            flash("❌ Email and password are required.")
            return redirect(url_for('admin_login'))

        try:
            # Fetch admin data
            result = supabase.table("admin_users").select("*").eq("email", email).execute()
            admin_data = result.data

            if not admin_data:
                flash("⚠️ No admin found with that email.")
                return redirect(url_for('admin_login'))

            admin = admin_data[0]

            # ✅ Compare password (plain text check or hashed check)
            if password != admin['password']:
                # Use check_password_hash(admin['password'], password) if hashed
                flash("❌ Incorrect password.")
                return redirect(url_for('admin_login'))

            # ✅ Login success
            session['admin'] = True
            session['admin_email'] = email
            session['admin_role'] = admin.get('role', 'admin')

            flash("✅ Login successful!")
            return redirect(url_for('admin_dashboard'))

        except Exception as e:
            flash(f"❌ Login failed: {e}")
            return redirect(url_for('admin_login'))

    return render_template("admin_login_password.html")

@app.route('/admin/manage-admins', methods=['GET', 'POST'])
def manage_admins():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    # Add new admin (POST)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        phone = request.form.get('phone', '').strip()
        password = request.form.get('password', '').strip()

        if not email or not password:
            flash("⚠️ Email and password are required.")
            return redirect(url_for('manage_admins'))

        try:
            # Check if admin already exists
            exists = supabase.table("admin_users").select("email").eq("email", email).execute()
            if exists.data:
                flash("⚠️ This email is already an admin.")
            else:
                hashed_password = generate_password_hash(password)
                supabase.table("admin_users").insert({
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "password": hashed_password
                }).execute()
                flash(f"✅ Admin '{email}' added successfully.")
        except Exception as e:
            flash(f"❌ Failed to add admin: {e}")

        return redirect(url_for('manage_admins'))

    # Display all admins (GET)
    try:
        admins = supabase.table("admin_users") \
                         .select("*") \
                         .order("created_at", desc=True) \
                         .execute().data
    except Exception as e:
        admins = []
        flash(f"❌ Failed to fetch admins: {e}")

    return render_template("manage_admins.html", admins=admins)

@app.route('/admin/add-admin', methods=['POST'])
def add_admin():
    if 'admin' not in session or session.get('admin_role') != 'superadmin':
        flash("Only superadmins can add new admins.")
        return redirect(url_for('admin_dashboard'))

    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip().lower()
    phone = request.form.get('phone', '').strip()
    password = request.form.get('password', '').strip()  # Optional if you want password too

    if not email or not password:
        flash("⚠️ Email and Password are required.")
        return redirect(url_for('manage_admins'))

    try:
        existing = supabase.table("admin_users").select("email").eq("email", email).execute()
        if existing.data:
            flash("⚠️ This email is already an admin.")
        else:
            supabase.table("admin_users").insert({
                "name": name,
                "email": email,
                "phone": phone,
                "password": password,
                "role": "admin",
                "created_at": datetime.utcnow().isoformat()  # Ensures created_at is filled
            }).execute()
            flash(f"✅ Admin '{email}' added.")
    except Exception as e:
        flash(f"❌ Failed to add admin: {e}")

    return redirect(url_for('manage_admins'))

@app.route('/admin/delete-admin', methods=['POST'])
def delete_admin():
    if 'admin' not in session or session.get('admin_role') != 'superadmin':
        flash("Only superadmins can delete admins.")
        return redirect(url_for('admin_dashboard'))

    email = request.form.get('email')
    if not email:
        flash("⚠️ Email is required.")
        return redirect(url_for('manage_admins'))

    try:
        supabase.table("admin_users").delete().eq("email", email).execute()
        flash(f"✅ Admin '{email}' deleted.")
    except Exception as e:
        flash(f"❌ Failed to delete admin: {e}")

    return redirect(url_for('manage_admins'))

#===========================
#Manage Basket Subjects
#===========================
@app.route('/admin/manage-basket', methods=['GET'])
def manage_basket():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    # Get filters from query parameters (if any)
    selected_program = request.args.get("program", "").strip()
    selected_branch = request.args.get("branch", "").strip()
    selected_basket = request.args.get("basket", "").strip()

    try:
        # Build Supabase filter query dynamically
        query = supabase.table("cbcs_basket").select("*")

        if selected_program:
            query = query.eq("program", selected_program)
        if selected_branch:
            query = query.eq("branch", selected_branch)
        if selected_basket:
            query = query.eq("basket", selected_basket)

        query = query.order("subject_code")
        subjects = query.execute().data

    except Exception as e:
        subjects = []
        flash(f"❌ Failed to load subjects: {e}")

    # Distinct values for dropdowns
    try:
        all_subjects = supabase.table("cbcs_basket").select("program, branch, basket").execute().data
        programs = sorted({row["program"] for row in all_subjects if row["program"]})
        branches = sorted({row["branch"] for row in all_subjects if row["branch"]})
        baskets = sorted({row["basket"] for row in all_subjects if row["basket"]})
    except:
        programs, branches, baskets = [], [], []

    return render_template("manage_basket.html",
                           subjects=subjects,
                           programs=programs,
                           branches=branches,
                           baskets=baskets,
                           selected_program=selected_program,
                           selected_branch=selected_branch,
                           selected_basket=selected_basket)

@app.route('/admin/add-subject', methods=['POST'])
def add_subject():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    # Extract and clean form data
    subject_code = request.form.get("subject_code", "").strip().upper()
    subject_name = request.form.get("subject_name", "").strip()
    credits = request.form.get("credits", "").strip()
    basket = request.form.get("basket", "").strip()
    program = request.form.get("program", "").strip()
    branch = request.form.get("branch", "").strip()

    # Basic validation
    if not subject_code or not subject_name or not credits or not program or not branch or not basket:
        flash("⚠️ All fields are required.")
        return redirect(url_for('manage_basket'))

    try:
        credits = float(credits)
    except ValueError:
        flash("⚠️ Credits must be a valid number.")
        return redirect(url_for('manage_basket'))

    try:
        # Check if the exact subject already exists
        existing = supabase.table("cbcs_basket").select("id").match({
            "subject_code": subject_code,
            "program": program,
            "branch": branch,
            "basket": basket
        }).execute()

        if existing.data:
            flash(f"⚠️ This subject already exists in the selected basket for this branch & program.")
        else:
            supabase.table("cbcs_basket").insert({
                "subject_code": subject_code,
                "subject_name": subject_name,
                "credits": credits,
                "basket": basket,
                "program": program,
                "branch": branch
            }).execute()
            flash(f"✅ Subject '{subject_code}' added.")
    except Exception as e:
        flash(f"❌ Failed to add subject: {e}")

    return redirect(url_for('manage_basket'))

@app.route('/admin/update-subject', methods=['POST'])
def update_subject():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    subject_code = request.form.get("subject_code", "").strip().upper()
    if not subject_code:
        flash("Subject code is required for update.")
        return redirect(url_for('manage_basket'))

    updates = {}
    if request.form.get("subject_name"):
        updates["subject_name"] = request.form["subject_name"].strip()
    if request.form.get("credits"):
        try:
            updates["credits"] = float(request.form["credits"])
        except:
            flash("Invalid credits.")
            return redirect(url_for('manage_basket'))
    if request.form.get("basket"):
        updates["basket"] = request.form["basket"].strip()

    if not updates:
        flash("No updates provided.")
        return redirect(url_for('manage_basket'))

    try:
        exists = supabase.table("cbcs_basket").select("id").eq("subject_code", subject_code).execute()
        if not exists.data:
            flash("❌ Subject not found.")
        else:
            subject_id = exists.data[0]["id"]
            supabase.table("cbcs_basket").update(updates).eq("id", subject_id).execute()
            flash(f"✅ Subject '{subject_code}' updated.")
    except Exception as e:
        flash(f"❌ Failed to update subject: {e}")

    return redirect(url_for('manage_basket'))

@app.route('/admin/upload-subjects', methods=['POST'])
def upload_subjects():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.")
        return redirect(url_for('manage_basket'))

    try:
        df = pd.read_excel(file)
        df.fillna('', inplace=True)

        inserted = 0
        skipped = 0
        for _, row in df.iterrows():
            subject_code = str(row.get("subject_code", "")).strip().upper()
            program = str(row.get("program", "")).strip()
            branch = str(row.get("branch", "")).strip()
            basket = str(row.get("basket", "")).strip()
            subject_name = str(row.get("subject_name", "")).strip()
            credits = str(row.get("credits", "")).strip()

            # Skip invalid
            if not all([subject_code, program, branch, basket, subject_name, credits]):
                skipped += 1
                continue

            # Check if subject already exists
            exists = supabase.table("cbcs_basket").select("id") \
                .eq("subject_code", subject_code) \
                .eq("program", program) \
                .eq("branch", branch).execute()

            if exists.data:
                skipped += 1
                continue

            supabase.table("cbcs_basket").insert({
                "subject_code": subject_code,
                "subject_name": subject_name,
                "credits": credits,
                "basket": basket,
                "program": program,
                "branch": branch
            }).execute()
            inserted += 1

        flash(f"✅ {inserted} subjects added. 🚫 {skipped} skipped (duplicates or incomplete).")
    except Exception as e:
        flash(f"❌ Upload failed: {e}")

    return redirect(url_for('manage_basket'))

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    session.pop('admin_email', None)
    session.pop('admin_role', None)
    flash("You have been logged out.")
    return redirect(url_for('admin_login'))

def generate_academic_years(start_year=2020):
    current_year = datetime.now().year
    # Include next academic year if past June
    if datetime.now().month > 6:
        current_year += 1
    return [f"{y}-{y + 1}" for y in range(start_year, current_year)]

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    # ==========================
    # School → Branch Mapping
    # ==========================
    school_branch_map = {
        "School of Engineering and Technology": [
            "Computer Science and Engineering",
            "Mechanical Engineering",
            "Electrical and Electronics Engineering",
            "Civil Engineering",
            "Agricultural Engineering",
            "Biotechnology Engineering",
            "Aerospace Engineering",
            "Bachelor of Computer Application (BCA)"
        ],
        "M.S. Swaminathan School of Agriculture": [
            "Agronomy", "Horticulture", "Soil Science"
        ],
        "School of Management": ["BBA", "MBA"],
        "School of Fisheries": ["Fisheries Science"],
        "School of Vocational Education and Training": ["Vocational Training"],
        "School of Applied Sciences": ["Physics", "Chemistry", "Mathematics"],
        "School of Agriculture & Bio-Engineering": ["Bio-Engineering"],
        "School of Veterinary and Animal Sciences": ["Veterinary Science"],
        "School of Nursing": ["Nursing"]
    }

    # ==========================
    # School → Program Mapping
    # ==========================
    school_program_map = {
        "School of Engineering and Technology": ["BTech", "BCA", "Diploma"],
        "M.S. Swaminathan School of Agriculture": ["BSc AG"],
        "School of Management": ["BBA", "MBA"],
        "School of Fisheries": ["BSc Fisheries"],
        "School of Vocational Education and Training": ["Diploma"],
        "School of Applied Sciences": ["BSc"],
        "School of Agriculture & Bio-Engineering": ["BTech Bio", "BSc Bio"],
        "School of Veterinary and Animal Sciences": ["BVSc"],
        "School of Nursing": ["BSc Nursing"]
    }

    # ✅ Academic years list
    academic_years = generate_academic_years()

    # ✅ Fetch uploaded result file records
    try:
        response = supabase.table("uploads").select("*").order("timestamp", desc=True).execute()
        uploads = response.data if response.data else []
    except Exception as e:
        uploads = []
        flash(f"Error fetching uploads: {e}")

    return render_template(
        "admin_dashboard.html",
        message=None,
        schoolBranchMap=school_branch_map,
        schoolProgramMap=school_program_map,
        uploads=uploads,
        academic_years=academic_years
    )

@app.route('/admin/upload_insert', methods=['POST'])
def upload_insert():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    file = request.files.get('file')
    school = request.form.get('school')
    branch = request.form.get('branch')
    semester = request.form.get('semester')
    academic_year = request.form.get('academic_year')
    program = request.form.get('program')

    if not file or not allowed_file(file.filename):
        flash("Please upload a valid .xlsx file.")
        return redirect(url_for('admin_dashboard'))

    filename = file.filename
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    df = pd.read_excel(filepath)
    df.fillna('', inplace=True)

    new_rows = []
    errors = []

    for _, row in df.iterrows():
        reg_no = str(row.get('Reg_No', '')).strip()
        subject_code = str(row.get('Subject_Code', '')).strip()

        if not reg_no or not subject_code:
            continue

        new_rows.append({
            "reg_no": reg_no,
            "name": str(row.get('Name', '')).strip(),
            "subject_code": subject_code,
            "subject_name": str(row.get('Subject_Name', '')).strip(),
            "type": str(row.get('Type', '')).strip(),
            "credits": str(row.get('Credits', '')).strip(),
            "grade": str(row.get('Grade', '')).strip(),
            "semester": semester,
            "school": school,
            "branch": branch,
            "academic_year": academic_year,
            "program": program
        })

    inserted = 0
    if new_rows:
        try:
            # You can consider chunking if the file is large
            supabase.table("results").insert(new_rows).execute()
            inserted = len(new_rows)
        except Exception as e:
            errors.append(f"Insertion failed: {e}")

    # Upload log
    try:
        supabase.table("uploads").insert({
            "filename": filename,
            "upload_type": "Sem Result",
            "school": school,
            "branch": branch,
            "semester": semester,
            "academic_year": academic_year,
            "program": program
        }).execute()
    except Exception as e:
        errors.append(f"Upload log failed: {e}")

    # Final flash messages
    if inserted:
        flash(f"✅ {inserted} rows inserted successfully.")
    else:
        flash("⚠️ No new records inserted.")

    if errors:
        flash("Some errors occurred: " + "; ".join(errors))

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/upload_update', methods=['POST'])
def upload_update():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    file = request.files.get('file')
    if not file or not allowed_file(file.filename):
        flash("Please upload a valid .xlsx file.")
        return redirect(url_for('admin_dashboard'))

    filename = file.filename
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        df = pd.read_excel(filepath)
        df.fillna('', inplace=True)
    except Exception as e:
        flash(f"Failed to read Excel file: {e}")
        return redirect(url_for('admin_dashboard'))

    # Rename columns
    df.columns = [col.strip().lower() for col in df.columns]
    column_map = {
        'registration no.': 'reg_no',
        'reg_no': 'reg_no',
        'subject code': 'subject_code',
        'grade': 'grade'
    }
    df.rename(columns={k: v for k, v in column_map.items() if k in df.columns}, inplace=True)

    if not {'reg_no', 'subject_code', 'grade'}.issubset(df.columns):
        flash("❌ Required columns missing: Registration No., Subject Code, Grade")
        return redirect(url_for('admin_dashboard'))

    # Build all (reg_no, subject_code) pairs from Excel
    keys = set()
    for _, row in df.iterrows():
        reg = str(row['reg_no']).strip()
        code = str(row['subject_code']).strip()
        if reg and code:
            keys.add((reg, code))

    if not keys:
        flash("❌ No valid records found.")
        return redirect(url_for('admin_dashboard'))

    # 🔍 Fetch all matching results in ONE query
    regnos = list({k[0] for k in keys})
    try:
        db_records = supabase.table("results").select("id, reg_no, subject_code") \
            .in_("reg_no", regnos).execute().data
    except Exception as e:
        flash(f"❌ Error fetching existing records: {e}")
        return redirect(url_for('admin_dashboard'))

    # 📌 Build lookup map
    id_map = {(r["reg_no"], r["subject_code"]): r["id"] for r in db_records}

    updated = 0
    errors = []

    for _, row in df.iterrows():
        reg_no = str(row.get("reg_no", "")).strip()
        subject_code = str(row.get("subject_code", "")).strip()
        grade = str(row.get("grade", "")).strip()

        key = (reg_no, subject_code)
        if key in id_map:
            try:
                supabase.table("results").update({
                    "grade": grade
                }).eq("id", id_map[key]).execute()
                updated += 1
            except Exception as e:
                errors.append(str(e))

    # ✅ Log the update
    try:
        supabase.table("uploads").insert({
            "filename": filename,
            "upload_type": "EOD"
        }).execute()
    except Exception as e:
        errors.append(f"Upload log failed: {e}")

    # ✅ Result messages
    if updated:
        flash(f"✅ {updated} grades updated successfully.")
    else:
        flash("⚠️ No matching records found to update.")

    if errors:
        flash("❌ Some errors occurred: " + "; ".join(errors))

    return redirect(url_for('admin_dashboard'))

@app.route('/get_semesters')
def get_semesters():
    reg_no = request.args.get('reg_no', '').strip()

    if not reg_no:
        return {'error': 'Registration number is required.', 'semesters': []}

    try:
        # Use `.select("semester").eq(...)` still due to Supabase limitation (no distinct in client lib)
        response = supabase.table("results") \
            .select("semester") \
            .eq("reg_no", reg_no) \
            .execute()

        # Collect and deduplicate semesters
        semesters = sorted(set(
            row["semester"] for row in response.data if row.get("semester")
        ))

        return {'semesters': semesters}

    except Exception as e:
        return {'error': str(e), 'semesters': []}

# ==========================
# Credit Tracker Logic and Routes
# ==========================

@app.route('/credit-tracker')
def credit_tracker_home():
    return render_template('credit_form.html')

@app.route('/view-credits')
def view_credits():
    reg_no = request.args.get('reg_no', '').strip()
    if not reg_no:
        return redirect(url_for('credit_tracker_home'))

    # ✅ Fetch all records of the student
    response = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    all_results = response.data

    if not all_results:
        return render_template("credit_view.html", student=None)

    # ✅ Extract student profile from the first result
    student = {
        "name": all_results[0].get("name", "-"),
        "reg_no": reg_no,
        "program": all_results[0].get("program", "-"),
        "school": all_results[0].get("school", "-"),
        "branch": all_results[0].get("branch", "-"),
        "batch": all_results[0].get("batch", "-"),
        "semester": all_results[-1].get("semester", "-"),
        "cgpa": "-",  # You can calculate if needed
    }

    # ✅ Define required credits based on program
    PROGRAM_CREDIT_REQUIREMENTS = {
        "BTech": 160,
        "Btech": 160,
        "BTech Honours": 180,
        "Btech Honours": 180,
        "BBA": 120,
        "Bba": 120,
        "BSc Ag": 172,
        "Bsc Ag": 172
    }

    program_key = student["program"].strip().title()
    required_credits = PROGRAM_CREDIT_REQUIREMENTS.get(program_key, 160)  # Default: 160

    # ✅ Initialize semester map and counters
    semester_map = {}
    cleared_credits = 0
    backlog_credits = 0
    backlogs = []

    for record in all_results:
        sem = record.get("semester", "Unknown")
        grade = record.get("grade", "").strip().upper()
        credit_str = record.get("credits", "0").strip()

        # ✅ Parse credits safely (e.g., handle "3+1")
        try:
            credits = sum(float(x) for x in credit_str.split('+'))
        except:
            credits = 0.0

        if sem not in semester_map:
            semester_map[sem] = {
                "semester": sem,
                "subjects": [],
                "total_credits": 0,
                "cleared_credits": 0,
                "backlog_credits": 0
            }

        semester_map[sem]["subjects"].append(record)
        semester_map[sem]["total_credits"] += credits

        if grade in ["F", "S"]:
            semester_map[sem]["backlog_credits"] += credits
            backlog_credits += credits
            backlogs.append(record)
        else:
            semester_map[sem]["cleared_credits"] += credits
            cleared_credits += credits

    # ✅ Final metrics
    completion_percentage = round((cleared_credits / required_credits) * 100, 2) if required_credits else 0

    return render_template(
        "credit_view.html",
        student=student,
        semester_data=list(semester_map.values()),
        credits_completed=cleared_credits,
        total_credits_required=required_credits,
        completion_percentage=completion_percentage,
        backlogs=backlogs,
        backlog_credits=backlog_credits
    )

@app.route('/download_semester_cards')
def download_semester_cards():
    reg_no = request.args.get('reg_no', '').strip()
    if not reg_no:
        return redirect(url_for('credit_tracker_home'))

    # 🔹 Fetch results
    response = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    all_results = response.data
    if not all_results:
        return "No results found", 404

    # 🔹 Student info
    student = {
        "name": all_results[0].get("name", "-"),
        "reg_no": reg_no,
        "program": all_results[0].get("program", "-"),
        "school": all_results[0].get("school", "-"),
        "branch": all_results[0].get("branch", "-"),
        "batch": all_results[0].get("batch", "-"),
        "semester": all_results[-1].get("semester", "-"),
        "cgpa": "-",
    }

    PROGRAM_CREDIT_REQUIREMENTS = {
        "BTech": 160, "Btech": 160,
        "BTech Honours": 180, "Btech Honours": 180,
        "BBA": 120, "Bba": 120,
        "BSc Ag": 172, "Bsc Ag": 172
    }
    program_key = student["program"].strip().title()
    required_credits = PROGRAM_CREDIT_REQUIREMENTS.get(program_key, 160)

    semester_map = {}
    cleared_credits = backlog_credits = 0
    backlog_subjects = []  # ✅ Collect all backlogs

    for record in all_results:
        sem = record.get("semester", "Unknown")
        grade = record.get("grade", "").strip().upper()
        credit_str = record.get("credits", "0").strip()

        try:
            credits = sum(float(x) for x in credit_str.split('+'))
        except:
            credits = 0.0

        if sem not in semester_map:
            semester_map[sem] = {
                "semester": sem,
                "subjects": [],
                "total_credits": 0,
                "cleared_credits": 0,
                "backlog_credits": 0
            }

        sub_info = {
            "subject_code": record.get("subject_code", "-"),
            "subject_name": record.get("subject_name", "-"),
            "type": record.get("type", "-"),
            "grade": grade,
            "credits": credit_str,
            "semester": sem
        }

        semester_map[sem]["subjects"].append(sub_info)
        semester_map[sem]["total_credits"] += credits

        if grade in ["F", "S"]:  # ✅ backlog condition
            semester_map[sem]["backlog_credits"] += credits
            backlog_credits += credits
            backlog_subjects.append(sub_info)
        else:
            semester_map[sem]["cleared_credits"] += credits
            cleared_credits += credits

    completion_percentage = round((cleared_credits / required_credits) * 100, 2) if required_credits else 0

    # 🔹 Generate PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    normal, bold = styles["Normal"], styles["Heading4"]

    elements.append(Paragraph(f"<b>Student Report Card</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    for key, label in [
        ("name", "Name"), ("reg_no", "Reg No"), ("school", "School"),
        ("program", "Program"), ("branch", "Branch"), ("batch", "Batch")
    ]:
        elements.append(Paragraph(f"<b>{label}:</b> {student[key]}", normal))
    elements.append(Spacer(1, 12))

    # 🔹 Summary Table
    summary_data = [["Total Required", "Completed", "Remaining", "Backlogs", "Completion %"],
        [required_credits, cleared_credits, required_credits - cleared_credits,
         backlog_credits, f"{completion_percentage}%"]]
    summary_table = Table(summary_data, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f1f1f1")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # 🔹 Semester-wise tables
    for sem in semester_map.values():
        elements.append(Paragraph(f"Semester {sem['semester']}", bold))
        elements.append(Spacer(1, 8))

        data = [["Code", "Name", "Type", "Grade", "Credits"]]
        for sub in sem["subjects"]:
            data.append([
                sub["subject_code"],
                Paragraph(sub["subject_name"], normal),
                sub["type"], sub["grade"], sub["credits"]
            ])

        table = Table(data, colWidths=[70, 200, 70, 60, 60], hAlign="LEFT")
        style = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#dbeeff")),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])

        # ✅ Highlight backlog rows in red
        for i, sub in enumerate(sem["subjects"], start=1):
            if sub["grade"] in ["F", "S"]:
                style.add("BACKGROUND", (0,i), (-1,i), colors.HexColor("#ffe5e5"))

        table.setStyle(style)
        elements.append(table)

        elements.append(Spacer(1, 6))
        elements.append(Paragraph(
            f"<b>Total:</b> {sem['total_credits']} | "
            f"<b>Cleared:</b> {sem['cleared_credits']} | "
            f"<b>Backlogs:</b> {sem['backlog_credits']}", normal
        ))
        elements.append(Spacer(1, 16))

    # 🔹 Overall Backlog Table (end of report)
    if backlog_subjects:
        elements.append(Paragraph("Overall Backlogs", bold))
        elements.append(Spacer(1, 8))

        backlog_data = [["Semester", "Code", "Name", "Grade", "Credits"]]
        for sub in backlog_subjects:
            backlog_data.append([
                sub["semester"], sub["subject_code"],
                Paragraph(sub["subject_name"], normal),
                sub["grade"], sub["credits"]
            ])

        backlog_table = Table(backlog_data, colWidths=[60, 70, 200, 60, 60], hAlign="LEFT")
        backlog_style = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#ffcccc")),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
        backlog_table.setStyle(backlog_style)
        elements.append(backlog_table)
        elements.append(Spacer(1, 16))

        # ✅ Add total backlog count
        elements.append(Paragraph(f"<b>Total Backlog Subjects:</b> {len(backlog_subjects)}", normal))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    return Response(pdf, mimetype='application/pdf',
        headers={"Content-Disposition": "attachment;filename=semester_report.pdf"})

@app.route('/download-credits-excel')
def download_semester_excel():
    reg_no = request.args.get('reg_no', '').strip()
    if not reg_no:
        return redirect(url_for('credit_tracker_home'))

    # ✅ Fetch all records of the student
    response = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    all_results = response.data

    if not all_results:
        return "No results found", 404

    # ✅ Extract student profile
    student = {
        "name": all_results[0].get("name", "-"),
        "reg_no": reg_no,
        "program": all_results[0].get("program", "-"),
        "school": all_results[0].get("school", "-"),
        "branch": all_results[0].get("branch", "-"),
        "batch": all_results[0].get("batch", "-"),
    }

    # ✅ Group into semesters + collect backlogs
    semester_map = {}
    backlog_subjects = []
    for record in all_results:
        sem = record.get("semester", "Unknown")
        grade = record.get("grade", "").strip().upper()
        if sem not in semester_map:
            semester_map[sem] = []
        semester_map[sem].append(record)

        if grade in ["F", "S"]:
            backlog_subjects.append(record)

    # ✅ Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Credit Report"

    # Column widths
    col_widths = [15, 50, 15, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Styles
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    backlog_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red

    row_num = 1

    # ✅ Student Info
    for field, value in student.items():
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=1, value=f"{field.title()}: {value}").font = bold_font
        ws.cell(row=row_num, column=1).alignment = center
        row_num += 1
    row_num += 1

    # ✅ Semester-wise data
    for sem, subjects in sorted(semester_map.items()):
        # Semester Heading
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=1, value=f"Semester {sem}").font = bold_font
        ws.cell(row=row_num, column=1).alignment = center
        row_num += 1

        # Table Header
        headers = ["Subject Code", "Subject Name", "Credits", "Grade"]
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=head)
            c.font = bold_font
            c.alignment = center
            c.fill = header_fill
            c.border = border
        row_num += 1

        # Table Rows
        for sub in subjects:
            subject_code = sub.get("subject_code", "-")
            subject_name = sub.get("subject_name", "-")
            credit_str = sub.get("credits", "0")
            grade = sub.get("grade", "-")

            for col, val in enumerate([subject_code, subject_name, credit_str, grade], 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.alignment = center
                c.border = border
                if grade in ["F", "S"]:  # highlight backlog subjects
                    c.fill = backlog_fill

            row_num += 1

        row_num += 2  # Gap before next semester

    # ✅ Overall Backlog Table
    if backlog_subjects:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=1, value="Overall Backlog Subjects").font = bold_font
        ws.cell(row=row_num, column=1).alignment = center
        row_num += 1

        headers = ["Subject Code", "Subject Name", "Credits", "Grade", "Semester"]
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=head)
            c.font = bold_font
            c.alignment = center
            c.fill = header_fill
            c.border = border
        row_num += 1

        for sub in backlog_subjects:
            subject_code = sub.get("subject_code", "-")
            subject_name = sub.get("subject_name", "-")
            credit_str = sub.get("credits", "0")
            grade = sub.get("grade", "-")
            sem = sub.get("semester", "-")

            for col, val in enumerate([subject_code, subject_name, credit_str, grade, sem], 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.alignment = center
                c.border = border
                c.fill = backlog_fill
            row_num += 1

    # ✅ Save to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"{student['reg_no']}_credits_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

#=============================
# Semester Wise Basket Analysis
#=============================

@app.route('/credit-report/<reg_no>')
def credit_report(reg_no):
    from flask import flash, redirect, url_for

    # --- Fetch student results ---
    response = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    results = response.data
    if not results:
        flash("No results found for this student", "danger")
        return redirect(url_for("index"))

    # --- Extract basic student info ---
    student_info = {
        "name": results[0].get("name", "-"),
        "reg_no": reg_no,
        "session": results[0].get("batch", "-"),
        "branch": results[0].get("branch", "-"),
        "section": results[0].get("section", "-"),
        "program": results[0].get("program", "BTech"),
        "total_backlog_credits": 0,
        "school": results[0].get("school", ""),   # ✅ new field
        "campus": results[0].get("campus", "")    # ✅ new field
    }

    # --- Default mapping if DB does not provide school/campus ---
    if not student_info["school"]:
        # Simple branch → school mapping
        branch = student_info["branch"].lower()
        if any(b in branch for b in ["cse", "civil", "eee", "mech", "ece", "it"]):
            student_info["school"] = "SCHOOL OF ENGINEERING & TECHNOLOGY"
        elif "bba" in branch or "mba" in branch:
            student_info["school"] = "SCHOOL OF MANAGEMENT"
        elif "ag" in branch or "agriculture" in branch:
            student_info["school"] = "SCHOOL OF AGRICULTURE & BIOSCIENCES"
        else:
            student_info["school"] = "SCHOOL OF STUDIES"

    if not student_info["campus"]:
        # You can expand this as needed
        student_info["campus"] = "PARALAKHEMUNDI CAMPUS"

    program = student_info["program"].strip().title()
    branch = student_info["branch"].strip().lower().replace(" ", "")

    # --- CBCS Basket Credit Requirements ---
    BASKET_CREDIT_REQUIREMENTS = {
        "Btech": {"Basket I": 17, "Basket II": 12, "Basket III": 25, "Basket IV": 58, "Basket V": 48, "Total": 160},
        "Btech Honours": {"Basket I": 17, "Basket II": 12, "Basket III": 25, "Basket IV": 58, "Basket V": 68, "Total": 180},
        "Bba": {"Basket I": 60, "Basket II": 32, "Basket III": 12, "Basket IV": 12, "Basket V": 4, "Total": 120},
        "Bsc Ag": {"Basket I": 18, "Basket II": 18, "Basket III": 20, "Basket IV": 96, "Basket V": 20, "Total": 172}
    }
    basket_requirements = BASKET_CREDIT_REQUIREMENTS.get(program, {})

    # --- Fetch CBCS basket mapping ---
    try:
        cbcs_resp = supabase.table("cbcs_basket").select("*").execute()
        cbcs_map = cbcs_resp.data
    except Exception as e:
        flash(f"Error fetching CBCS mappings: {e}", "danger")
        cbcs_map = []

    # --- Build subject → basket mapping ---
    subject_basket_map = {}
    basket_indices = {"Basket I": 0, "Basket II": 1, "Basket III": 2, "Basket IV": 3, "Basket V": 4}
    for row in cbcs_map:
        sub_code = row.get("subject_code", "").strip()
        cbcs_prog = row.get("program", "").strip().title()
        cbcs_branch = row.get("branch", "").strip().lower().replace(" ", "")

        match = (
            (cbcs_prog == program and cbcs_branch == branch) or
            (cbcs_prog == program and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == branch)
        )

        if sub_code and match:
            try:
                credits = float(row.get("credits", 0))
            except:
                credits = 0
            subject_basket_map[sub_code] = {
                "basket": row.get("basket", "Unknown"),
                "credits": credits,
                "subject_name": row.get("subject_name", "")
            }

    # --- Group results into semesters and baskets ---
    semesters, totals, backlog, baskets = {}, {}, {}, {}
    unmatched = 0

    for r in results:
        year = r.get("year", "-")
        semester = r.get("semester", "-").strip().title().replace("Sem", "Semester-").replace(" ", "-")
        sub_code = r.get("subject_code", "").strip()
        grade = r.get("grade", "").strip().upper()

        semesters.setdefault(year, {}).setdefault(semester, [])
        totals.setdefault(year, {}).setdefault(semester, None)
        backlog.setdefault(year, {}).setdefault(semester, [])

        if sub_code in subject_basket_map:
            info = subject_basket_map[sub_code]
            basket_name = info["basket"]
            credit = info["credits"]
            sub_name = r.get("subject_name") or info["subject_name"]

            if basket_name not in baskets:
                baskets[basket_name] = {
                    "completed": [],
                    "backlogs": [],
                    "total_credits": basket_requirements.get(basket_name, 0),
                    "cleared_credits": 0,
                    "backlog_credits": 0
                }

            sub_data = {
                "subject_code": sub_code,
                "subject_name": sub_name,
                "credits": credit,
                "grade": grade,
                "semester": semester
            }

            if grade in ["F", "BACKLOG", "S"]:
                baskets[basket_name]["backlogs"].append(sub_data)
                baskets[basket_name]["backlog_credits"] += credit
                student_info["total_backlog_credits"] += credit
            else:
                baskets[basket_name]["completed"].append(sub_data)
                baskets[basket_name]["cleared_credits"] += credit

            basket_cols = [0, 0, 0, 0, 0]
            idx = basket_indices.get(basket_name, None)
            if idx is not None:
                basket_cols[idx] = credit

            semesters[year][semester].append({
                "sl_no": len(semesters[year][semester]) + 1,
                "code": sub_code,
                "subject": sub_name,
                "baskets": basket_cols,
                "grade": grade
            })
        else:
            unmatched += 1

    if unmatched > 0:
        flash(f"{unmatched} subject(s) did not match any CBCS basket mapping.", "info")

    # --- Compute totals row per semester ---
    for year, sems in semesters.items():
        for sem_name, subjects in sems.items():
            basket_totals = [0, 0, 0, 0, 0]
            for sub in subjects:
                for i in range(5):
                    basket_totals[i] += sub["baskets"][i]
            grand_total = sum(basket_totals)
            totals[year][sem_name] = {
                "sl_no": "",
                "code": "",
                "subject": "Total",
                "baskets": basket_totals,
                "grand_total": grand_total,
                "grade": ""
            }

    # --- Compute year-level totals ---
    year_totals = {}
    for year, sems in totals.items():
        sem_names = sorted(sems.keys())
        year_totals[year] = []
        temp_totals = [0, 0, 0, 0, 0]
        temp_grand = 0
        for idx, sem_name in enumerate(sem_names, start=1):
            for i in range(5):
                temp_totals[i] += totals[year][sem_name]["baskets"][i]
            temp_grand += totals[year][sem_name]["grand_total"]
            if idx % 2 == 0:
                year_totals[year].append({
                    "sl_no": "",
                    "code": "",
                    "subject": f"{year} Total Credits",
                    "baskets": temp_totals.copy(),
                    "grand_total": temp_grand,
                    "grade": ""
                })
                temp_totals = [0, 0, 0, 0, 0]
                temp_grand = 0

    # --- Prepare semester-wise summary ---
    semester_summary = []
    basket_sum = [0, 0, 0, 0, 0]
    grand_sum = 0
    sem_counter = 1

    for year, sems in sorted(semesters.items()):
        for sem_name, subjects in sorted(sems.items()):
            basket_totals = [0, 0, 0, 0, 0]
            for sub in subjects:
                for i in range(5):
                    basket_totals[i] += sub["baskets"][i]
            sem_grand = sum(basket_totals)
            semester_summary.append({
                "sem_number": sem_counter,
                "baskets": basket_totals,
                "grand_total": sem_grand
            })
            for i in range(5):
                basket_sum[i] += basket_totals[i]
            grand_sum += sem_grand
            sem_counter += 1

    # --- Sort baskets ---
    def roman_to_int(roman):
        return {"I":1,"II":2,"III":3,"IV":4,"V":5}.get(roman.upper(),None)

    def basket_order_key(basket_name):
        parts = basket_name.strip().split()
        if len(parts) > 1:
            roman = parts[-1]
            if roman.isdigit(): return int(roman)
            val = roman_to_int(roman)
            if val: return val
        return 999

    sorted_baskets = dict(sorted(baskets.items(), key=lambda x: basket_order_key(x[0])))

    # --- Prepare data for template ---
    data = {
        "student_info": student_info,
        "semesters": semesters,
        "totals": totals,
        "year_totals": year_totals,
        "backlog": backlog,
        "baskets": sorted_baskets,
        "basket_requirements": basket_requirements,
        "semester_summary": semester_summary,
        "basket_sum": basket_sum,
        "grand_sum": grand_sum
    }

    return render_template("credit_report.html", data=data)

#=============================
# Basket Summary Logic and Route
#=============================

@app.route('/view-basket-subjects')
def view_basket_subjects():
    reg_no = request.args.get('reg_no', '').strip()
    if not reg_no:
        return redirect(url_for('credit_tracker_home'))

    # ✅ Fetch student results
    results_resp = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    results = results_resp.data

    if not results:
        return render_template("basket_subjects.html", student=None)

    # ✅ Extract student info
    student = {
        "name": results[0].get("name", "-"),
        "reg_no": reg_no,
        "school": results[0].get("school", "-"),
        "branch": results[0].get("branch", "-"),
        "program": results[0].get("program", "-"),
        "batch": results[0].get("batch", "-")
    }

    # ✅ Normalize
    program = student["program"].strip().title() if student["program"] else "BTech"
    branch = student["branch"].strip().lower().replace(" ", "") if student["branch"] else "all"

    # ✅ Basket Credit Requirements
    BASKET_CREDIT_REQUIREMENTS = {
        "Btech": {
            "Basket I": 17, "Basket II": 12, "Basket III": 25,
            "Basket IV": 58, "Basket V": 48, "Total": 160
        },
        "Btech Honours": {
            "Basket I": 17, "Basket II": 12, "Basket III": 25,
            "Basket IV": 58, "Basket V": 68, "Total": 180
        },
        "Bba": {
            "Basket I": 60, "Basket II": 32, "Basket III": 12,
            "Basket IV": 12, "Basket V": 4, "Total": 120
        },
        "Bsc Ag": {
            "Basket I": 18, "Basket II": 18, "Basket III": 20,
            "Basket IV": 96, "Basket V": 20, "Total": 172
        }
    }

    basket_requirements = BASKET_CREDIT_REQUIREMENTS.get(program.title(), {})

    # ✅ Fetch CBCS data
    try:
        cbcs_resp = supabase.table("cbcs_basket").select("*").execute()
        cbcs_map = cbcs_resp.data
    except Exception as e:
        flash(f"❌ Error fetching CBCS mappings: {e}", "danger")
        return render_template("basket_subjects.html", student=student, basket_data={}, basket_requirements={})

    # ✅ Build subject_code → basket map with filtering
    subject_basket_map = {}
    for row in cbcs_map:
        sub_code = row.get("subject_code", "").strip()
        cbcs_prog = row.get("program", "").strip().title()
        cbcs_branch = row.get("branch", "").strip().lower().replace(" ", "")

        # Match combinations
        match = (
            (cbcs_prog == program and cbcs_branch == branch) or
            (cbcs_prog == program and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == branch)
        )

        if sub_code and match:
            try:
                credits = float(row.get("credits", 0))
            except:
                credits = 0
            subject_basket_map[sub_code] = {
                "basket": row.get("basket", "Unknown"),
                "credits": credits,
                "subject_name": row.get("subject_name", "")
            }

    # ✅ Group into baskets
    baskets = {}
    unmatched = 0

    for row in results:
        sub_code = row.get("subject_code", "").strip()
        grade = row.get("grade", "").strip().upper()

        if sub_code not in subject_basket_map:
            unmatched += 1
            continue

        info = subject_basket_map[sub_code]
        basket = info["basket"]
        credit = info["credits"]
        sub_name = row.get("subject_name") or info["subject_name"]

        if basket not in baskets:
            baskets[basket] = {
                "completed": [],
                "backlogs": [],
                "total_credits": basket_requirements.get(basket, 0),
                "cleared_credits": 0,
                "backlog_credits": 0
            }

        sub_data = {
            "subject_code": sub_code,
            "subject_name": sub_name,
            "credits": credit,
            "grade": grade,
            "semester": row.get("semester", "-")
        }

        if grade in ["F", "S"]:
            baskets[basket]["backlogs"].append(sub_data)
            baskets[basket]["backlog_credits"] += credit
        else:
            baskets[basket]["completed"].append(sub_data)
            baskets[basket]["cleared_credits"] += credit

    if unmatched > 0:
        flash(f"{unmatched} subject(s) did not match any CBCS basket mapping and were ignored.", "info")
    
    # ✅ Roman numeral to integer helper (properly indented inside function)
    def roman_to_int(roman):
        roman_map = {
            "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
            "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10
        }
        return roman_map.get(roman.upper(), None)

    # ✅ Sort baskets (works for both Roman & numeric)
    def basket_order_key(basket_name):
        parts = basket_name.strip().split()
        if len(parts) > 1:
            roman_or_num = parts[-1]
            if roman_or_num.isdigit():
                return int(roman_or_num)
            val = roman_to_int(roman_or_num)
            if val:
                return val
        return 999

    sorted_baskets = dict(sorted(baskets.items(), key=lambda x: basket_order_key(x[0])))

    return render_template(
        "basket_subjects.html",
        student=student,
        basket_data=sorted_baskets,
        basket_requirements=basket_requirements
    )

@app.route("/download_subject_excel/<reg_no>")
def download_subject_excel(reg_no):
    import io, re, xlsxwriter
    from flask import send_file

    # 🔹 Fetch student results
    results_resp = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    results = results_resp.data
    if not results:
        return "No data to export", 400

    # 🔹 Extract student info
    student = {
        "name": results[0].get("name", "-"),
        "reg_no": reg_no,
        "school": results[0].get("school", "-"),
        "branch": results[0].get("branch", "-"),
        "program": results[0].get("program", "-"),
        "batch": results[0].get("batch", "-"),
    }

    # ✅ Normalize program/branch
    program = student["program"].strip().title() if student["program"] else "BTech"
    branch = student["branch"].strip().lower().replace(" ", "") if student["branch"] else "all"

    # ✅ CBCS basket mappings
    cbcs_resp = supabase.table("cbcs_basket").select("*").execute()
    cbcs_map = cbcs_resp.data
    subject_basket_map = {}
    for row in cbcs_map:
        sub_code = row.get("subject_code", "").strip()
        cbcs_prog = row.get("program", "").strip().title()
        cbcs_branch = row.get("branch", "").strip().lower().replace(" ", "")
        match = (
            (cbcs_prog == program and cbcs_branch == branch) or
            (cbcs_prog == program and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == branch)
        )
        if sub_code and match:
            subject_basket_map[sub_code] = row.get("basket", "Unknown")

    # ✅ Build baskets
    baskets = {}
    all_backlogs = []
    for row in results:
        sub_code = row.get("subject_code", "").strip()
        basket = subject_basket_map.get(sub_code, "Unknown")
        sub_data = {
            "basket": basket,
            "subject_code": row.get("subject_code"),
            "subject_name": row.get("subject_name"),
            "credits": row.get("credits"),
            "grade": row.get("grade"),
            "semester": row.get("semester"),
        }
        baskets.setdefault(basket, []).append(sub_data)
        if row.get("grade", "").upper() in ["F", "S"]:
            all_backlogs.append(sub_data)

    # ✅ Helpers
    def roman_to_int(roman):
        roman_map = {"I":1,"II":2,"III":3,"IV":4,"V":5,"VI":6,"VII":7,"VIII":8,"IX":9,"X":10}
        return roman_map.get(roman.upper(), None)

    def basket_order_key(basket_name):
        match = re.search(r'(?:Basket\s*)([IVXLCDM]+|\d+)$', basket_name.strip(), re.IGNORECASE)
        if match:
            val = match.group(1)
            if val.isdigit(): return int(val)
            roman_val = roman_to_int(val.upper())
            if roman_val: return roman_val
        return 0

    # 🔹 Generate Excel
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet("Subjects")

    # Formats
    header_fmt = workbook.add_format({"bold": True, "bg_color": "#ADD8E6", "border": 1, "align": "center", "valign": "vcenter"})
    backlog_fmt = workbook.add_format({"bg_color": "#FDE2E2", "border": 1, "text_wrap": True, "align": "center", "valign": "vcenter"})
    normal_fmt = workbook.add_format({"border": 1, "text_wrap": True, "align": "center", "valign": "vcenter"})
    title_fmt = workbook.add_format({"bold": True, "font_size": 14, "align": "center"})
    subheader_fmt = workbook.add_format({"bold": True, "font_color": "blue", "underline": 1, "align": "center"})
    student_info_fmt = workbook.add_format({"bold": True, "bg_color": "#E8F4F8", "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True})

    row = 0
    worksheet.merge_range(row, 0, row, 5, f"Subject Report - {student['reg_no']}", title_fmt)
    row += 2

    # Merge student info cells across all columns and center
    info_labels = ["Name", "Reg No", "School", "Branch", "Program", "Batch"]
    for label in info_labels:
        value = student[label.lower().replace(" ", "_")]
        cell_text = f"{label}: {value}"
        worksheet.merge_range(row, 0, row, 5, cell_text, student_info_fmt)
        # dynamic row height
        row_height = max(15, min(len(cell_text)//30 * 15 + 15, 45))
        worksheet.set_row(row, row_height)
        row += 1
    row += 1

    # Column headers
    headers = ["Basket", "Code", "Name", "Credits", "Grade", "Sem"]

    # Iterate sorted baskets
    for basket, subs in sorted(baskets.items(), key=lambda x: basket_order_key(x[0])):
        worksheet.merge_range(row, 0, row, 5, f"{basket} - Subjects", subheader_fmt)
        row += 1
        for col, h in enumerate(headers):
            worksheet.write(row, col, h, header_fmt)
        row += 1
        subs_sorted = sorted(subs, key=lambda x: (str(x.get("semester", "")), x.get("subject_code", "")))
        for sub in subs_sorted:
            fmt = backlog_fmt if sub["grade"].upper() in ["F", "S"] else normal_fmt
            worksheet.write(row, 0, sub["basket"], fmt)
            worksheet.write(row, 1, sub["subject_code"], fmt)
            worksheet.write(row, 2, sub["subject_name"], fmt)
            worksheet.write(row, 3, sub["credits"], fmt)
            worksheet.write(row, 4, sub["grade"], fmt)
            worksheet.write(row, 5, sub["semester"], fmt)
            row += 1
        row += 2

    # Overall Backlog Summary
    if all_backlogs:
        worksheet.merge_range(row, 0, row, 5, "Overall Backlog Summary", subheader_fmt)
        row += 1
        for col, h in enumerate(headers):
            worksheet.write(row, col, h, header_fmt)
        row += 1
        for sub in all_backlogs:
            worksheet.write(row, 0, sub["basket"], backlog_fmt)
            worksheet.write(row, 1, sub["subject_code"], backlog_fmt)
            worksheet.write(row, 2, sub["subject_name"], backlog_fmt)
            worksheet.write(row, 3, sub["credits"], backlog_fmt)
            worksheet.write(row, 4, sub["grade"], backlog_fmt)
            worksheet.write(row, 5, sub["semester"], backlog_fmt)
            row += 1

    # Adjust column widths
    worksheet.set_column("A:A", 12)
    worksheet.set_column("B:B", 18)
    worksheet.set_column("C:C", 50)
    worksheet.set_column("D:D", 10)
    worksheet.set_column("E:E", 10)
    worksheet.set_column("F:F", 8)

    workbook.close()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"{reg_no}_subjects.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/download_subject_pdf/<reg_no>")
def download_subject_pdf(reg_no):

    # 🔹 Fetch student results
    results_resp = supabase.table("results").select("*").eq("reg_no", reg_no).execute()
    results = results_resp.data
    if not results:
        return "No data to export", 400

    # 🔹 Extract student info
    student = {
        "name": results[0].get("name", "-"),
        "reg_no": reg_no,
        "school": results[0].get("school", "-"),
        "branch": results[0].get("branch", "-"),
        "program": results[0].get("program", "-"),
        "batch": results[0].get("batch", "-"),
    }

    # ✅ Normalize program/branch
    program = student["program"].strip().title() if student["program"] else "BTech"
    branch = student["branch"].strip().lower().replace(" ", "") if student["branch"] else "all"

    # ✅ CBCS basket mappings
    cbcs_resp = supabase.table("cbcs_basket").select("*").execute()
    cbcs_map = cbcs_resp.data

    subject_basket_map = {}
    for row in cbcs_map:
        sub_code = row.get("subject_code", "").strip()
        cbcs_prog = row.get("program", "").strip().title()
        cbcs_branch = row.get("branch", "").strip().lower().replace(" ", "")
        match = (
            (cbcs_prog == program and cbcs_branch == branch) or
            (cbcs_prog == program and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == "all") or
            (cbcs_prog == "All" and cbcs_branch == branch)
        )
        if sub_code and match:
            subject_basket_map[sub_code] = row.get("basket", "Unknown")

    # ✅ Build baskets
    baskets = {}
    all_backlogs = []
    for row in results:
        sub_code = row.get("subject_code", "").strip()
        basket = subject_basket_map.get(sub_code, "Unknown")
        sub_data = {
            "basket": basket,
            "subject_code": row.get("subject_code"),
            "subject_name": row.get("subject_name"),
            "credits": row.get("credits"),
            "grade": row.get("grade"),
            "semester": row.get("semester"),
        }
        baskets.setdefault(basket, []).append(sub_data)
        if row.get("grade", "").upper() in ["F", "S"]:
            all_backlogs.append(sub_data)

    # ✅ Helpers for basket sorting
    def roman_to_int(roman):
        roman_map = {"I":1,"II":2,"III":3,"IV":4,"V":5,"VI":6,"VII":7,"VIII":8,"IX":9,"X":10}
        return roman_map.get(roman.upper(), None)

    def basket_order_key(basket_name):
        match = re.search(r'(?:Basket\s*)([IVXLCDM]+|\d+)$', basket_name.strip(), re.IGNORECASE)
        if match:
            val = match.group(1)
            if val.isdigit(): return int(val)
            roman_val = roman_to_int(val.upper())
            if roman_val: return roman_val
        return 0

    # 🔹 PDF Setup
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]

    # Title
    elements.append(Paragraph(f"Subject Report - {student['reg_no']}", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Student Info
    student_info = f"""
    <b>Name:</b> {student['name']}<br/>
    <b>Reg No:</b> {student['reg_no']}<br/>
    <b>School:</b> {student['school']}<br/>
    <b>Branch:</b> {student['branch']}<br/>
    <b>Program:</b> {student['program']}<br/>
    <b>Batch:</b> {student['batch']}<br/>
    """
    elements.append(Paragraph(student_info, styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Column widths (name column wider for wrapping)
    col_widths = [70, 70, 220, 60, 50, 40]

    def wrap_row(row):
        """Convert strings into Paragraph for wrapping"""
        wrapped = []
        for item in row:
            if isinstance(item, str):
                wrapped.append(Paragraph(item, normal_style))
            else:
                wrapped.append(item)
        return wrapped

    # Basket-wise subjects (sorted)
    for basket, subs in sorted(baskets.items(), key=lambda x: basket_order_key(x[0])):
        elements.append(Paragraph(f"{basket} - Subjects", styles["Heading3"]))

        table_data = [["Basket", "Code", "Name", "Credits", "Grade", "Sem"]]
        row_styles = []

        # Sort subjects by semester then code
        subs_sorted = sorted(subs, key=lambda x: (str(x.get("semester", "")), x.get("subject_code", "")))

        for i, sub in enumerate(subs_sorted, start=1):
            row = [
                sub["basket"], sub["subject_code"], sub["subject_name"],
                sub["credits"], sub["grade"], sub["semester"]
            ]
            table_data.append(row)
            if sub["grade"].upper() in ["F", "S"]:
                row_styles.append(("BACKGROUND", (0, i), (-1, i), colors.lavenderblush))

        # Wrap text
        table_data = [wrap_row(r) for r in table_data]

        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ] + row_styles))

        elements.append(table)
        elements.append(Spacer(1, 18))

    # Overall Backlog Summary
    if all_backlogs:
        elements.append(Paragraph("Overall Backlog Summary", styles["Heading2"]))
        summary_data = [["Basket", "Code", "Name", "Credits", "Grade", "Sem"]]
        for sub in all_backlogs:
            summary_data.append([sub["basket"], sub["subject_code"], sub["subject_name"],
                                 sub["credits"], sub["grade"], sub["semester"]])
        summary_data = [wrap_row(r) for r in summary_data]

        table = Table(summary_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.red),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)

    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name=f"{reg_no}_subjects.pdf",
        mimetype="application/pdf"
    )

@app.route('/basket-summary-report', methods=['GET', 'POST'])
def basket_summary_report():
    try:
        # 📌 Dropdown values from results table
        dropdown_data = supabase.table("results").select("school", "branch", "program").execute().data
        unique_schools = sorted({d.get("school") for d in dropdown_data if d.get("school")})
        unique_programs = sorted({d.get("program") for d in dropdown_data if d.get("program")})
        unique_branches = sorted({d.get("branch") for d in dropdown_data if d.get("branch")})

    except Exception as e:
        flash(f"❌ Error fetching dropdowns: {e}", "danger")
        return render_template("basket_summary_report.html", data=[], baskets=[], schools=[], programs=[], branches=[])

    if request.method == 'POST':
        school = request.form.get('school')
        program = request.form.get('program')
        branch = request.form.get('branch')
        start_reg = request.form.get('start_reg')
        end_reg = request.form.get('end_reg')

        try:
            # ✅ Get student results in range
            student_resp = supabase.table("results") \
                .select("*") \
                .gte("reg_no", start_reg) \
                .lte("reg_no", end_reg) \
                .eq("school", school) \
                .eq("branch", branch) \
                .eq("program", program) \
                .execute()

            student_results = student_resp.data
            if not student_results:
                flash("⚠️ No student records found.", "warning")
                return render_template("basket_summary_report.html", data=[], baskets=[], schools=unique_schools, programs=unique_programs, branches=unique_branches)

            # ✅ CBCS Data
            cbcs_map = supabase.table("cbcs_basket").select("*").execute().data

            # 🔁 Build subject_code → list of basket options
            subject_basket_map = defaultdict(list)
            for row in cbcs_map:
                sub_code = row.get("subject_code", "").strip()
                prog = row.get("program", "").strip().lower()
                brnch = row.get("branch", "").strip().lower().replace(" ", "")
                basket = row.get("basket", "").strip()
                credits = float(row.get("credits", 0))

                if sub_code and basket:
                    subject_basket_map[sub_code].append({
                        "program": prog,
                        "branch": brnch,
                        "basket": basket,
                        "credits": credits
                    })

            # ✅ Basket Requirement Map
            BASKET_CREDIT_REQUIREMENTS = {
                "Btech": {
                    "Basket I": 17, "Basket II": 12, "Basket III": 25,
                    "Basket IV": 58, "Basket V": 48, "Total": 160
                },
                "Btech Honours": {
                    "Basket I": 17, "Basket II": 12, "Basket III": 25,
                    "Basket IV": 58, "Basket V": 68, "Total": 180
                },
                "Bba": {
                    "Basket I": 60, "Basket II": 32, "Basket III": 12,
                    "Basket IV": 12, "Basket V": 4, "Total": 120
                },
                "Bsc Ag": {
                    "Basket I": 18, "Basket II": 18, "Basket III": 20,
                    "Basket IV": 96, "Basket V": 20, "Total": 172
                }
            }
            basket_labels = list(BASKET_CREDIT_REQUIREMENTS.get(program.title(), {}).keys())[:-1]  # Exclude 'Total'

            # ✅ Aggregation per student
            student_data = defaultdict(lambda: {
                "name": "",
                "reg_no": "",
                "branch": "",
                "baskets": defaultdict(float),
                "backlog_credits": 0.0,
                "total": 0.0
            })

            # ✅ Track only the latest attempt per subject per student
            student_subjects = defaultdict(dict)

            for row in student_results:
                reg_no = row.get("reg_no", "").strip()
                name = row.get("name", "-")
                br = row.get("branch", "-")
                subject_code = row.get("subject_code", "").strip()
                grade = row.get("grade", "").strip().upper()
                semester = int(row.get("semester", 0) or 0)

                student = student_data[reg_no]
                student["name"] = name
                student["reg_no"] = reg_no
                student["branch"] = br

                if subject_code not in subject_basket_map:
                    continue

                matched = None
                for option in subject_basket_map[subject_code]:
                    prog = option["program"]
                    brnch = option["branch"]
                    if (prog == program.lower() or prog == "all") and (brnch == branch.lower().replace(" ", "") or brnch == "all"):
                        matched = option
                        break

                if not matched:
                    continue

                credits = matched["credits"]
                basket = matched["basket"]

                # ✅ Keep only latest attempt
                prev = student_subjects[reg_no].get(subject_code)
                if not prev or semester > prev["semester"]:
                    student_subjects[reg_no][subject_code] = {
                        "grade": grade,
                        "credits": credits,
                        "basket": basket,
                        "semester": semester
                    }

            # ✅ Now aggregate from deduped subjects
            for reg_no, subjects in student_subjects.items():
                student = student_data[reg_no]
                for subject_code, info in subjects.items():
                    if info["grade"] in ["F", "S"]:
                        student["backlog_credits"] += info["credits"]
                    else:
                        student["baskets"][info["basket"]] += info["credits"]
                        student["total"] += info["credits"]

            final_data = sorted(student_data.values(), key=lambda x: x["reg_no"])

            # ✅ Save in session for downloads
            session["basket_summary_data"] = final_data
            session["basket_labels"] = basket_labels

            return render_template("basket_summary_report.html",
                                   data=final_data,
                                   baskets=basket_labels,
                                   schools=unique_schools,
                                   programs=unique_programs,
                                   branches=unique_branches)

        except Exception as e:
            flash(f"❌ Error processing report: {str(e)}", "danger")
            return render_template("basket_summary_report.html", data=[], baskets=[], schools=unique_schools, programs=unique_programs, branches=unique_branches)

    # GET method → clear old session cache
    session.pop("basket_summary_data", None)
    session.pop("basket_labels", None)
    return render_template("basket_summary_report.html",
                           data=[],
                           baskets=[],
                           schools=unique_schools,
                           programs=unique_programs,
                           branches=unique_branches)

@app.route("/download_basket_excel")
def download_basket_excel():
    data = session.get("basket_summary_data", [])
    baskets = session.get("basket_labels", [])

    if not data:
        return "No data to export", 400

    # Convert to DataFrame
    rows = []
    for idx, student in enumerate(data, 1):
        row = {
            "Sl.No": idx,
            "Name": student["name"],
            "Registration No": student["reg_no"],
            "Department": student["branch"],
        }
        for basket in baskets:
            row[basket] = student["baskets"].get(basket, 0)
        row["Total Credits"] = student["total"]
        row["Backlog Credits"] = student["backlog_credits"]
        rows.append(row)

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Basket Summary")

    output.seek(0)
    response = make_response(output.read())
    response.headers["Content-Disposition"] = "attachment; filename=basket_summary.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route("/download_basket_pdf")
def download_basket_pdf():
    data = session.get("basket_summary_data", [])
    baskets = session.get("basket_labels", [])

    if not data:
        return "No data to export", 400

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Build table header
    header = ["Sl.No", "Name", "Registration No", "Department"] + baskets + ["Total Credits", "Backlog Credits"]

    # Build rows
    table_data = [header]
    for idx, student in enumerate(data, 1):
        row = [
            idx,
            student["name"],
            student["reg_no"],
            student["branch"],
        ]
        for basket in baskets:
            row.append(student["baskets"].get(basket, 0))
        row.append(student["total"])
        row.append(student["backlog_credits"])
        table_data.append(row)

    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
    ]))

    elements.append(Paragraph("📊 Basket Summary Report", styles["Title"]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers["Content-Disposition"] = "attachment; filename=basket_summary.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return response

#==========================
#=== Logic for Sitemap ===
#==========================

@app.route('/sitemap.xml')
def sitemap():
    base_url = "https://my-result.onrender.com"

    # Static pages to include
    static_routes = [
        "/", "/credit-tracker", "/topper", "/basket-summary-report"
    ]

    # Fetch distinct reg_nos from Supabase
    try:
        response = supabase.table("results").select("reg_no").execute()
        all_reg_nos = list({row["reg_no"] for row in response.data if "reg_no" in row})
    except Exception as e:
        print("Supabase error in sitemap:", e)
        all_reg_nos = []

    # Start XML
    xml = ['<?xml version="1.0" encoding="UTF-8"?>']
    xml.append('<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">')

    # Static URLs
    for route in static_routes:
        xml.append("  <url>")
        xml.append(f"    <loc>{base_url}{route}</loc>")
        xml.append(f"    <lastmod>{datetime.utcnow().date()}</lastmod>")
        xml.append("    <changefreq>weekly</changefreq>")
        xml.append("    <priority>0.8</priority>")
        xml.append("  </url>")

    # Dynamic URLs for each reg_no
    for reg_no in all_reg_nos:
        for route in ["view-credits", "view-basket-subjects", "result"]:
            xml.append("  <url>")
            xml.append(f"    <loc>{base_url}/{route}?reg_no={reg_no}</loc>")
            xml.append(f"    <lastmod>{datetime.utcnow().date()}</lastmod>")
            xml.append("    <changefreq>weekly</changefreq>")
            xml.append("    <priority>0.6</priority>")
            xml.append("  </url>")

    xml.append("</urlset>")
    return Response("\n".join(xml), mimetype="application/xml")

@app.route('/robots.txt')
def robots_txt():
    return Response(
        "User-agent: *\n"
        "Disallow: /admin/\n"
        "Disallow: /uploads/\n"
        "Disallow: /static/\n"
        "Allow: /\n"
        "Sitemap: https://my-result.onrender.com/sitemap.xml",
        mimetype="text/plain"
    )

@app.route("/about")
def about():
    return render_template("about.html")

@app.route('/google53dcc92479ba04f1.html')
def google_verify():
    return 'google-site-verification: google53dcc92479ba04f1.html'

if __name__ == '__main__':
    app.run(debug=True)
